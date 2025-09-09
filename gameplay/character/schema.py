# gameplay/character/schema.py
from __future__ import annotations
import re
from dataclasses import dataclass
from typing import Dict, List
from openpyxl import load_workbook
from core.settings import load_settings
from core.config import DATA_DIR

XLSX_PATH = DATA_DIR / "creation_player.xlsx"

ATTR_MAP = {
    "força": "STR", "forca": "STR",
    "destreza": "DEX",
    "inteligência": "INT", "inteligencia": "INT",
    "constituição": "CON", "constituicao": "CON",
}

def _slug(s: str) -> str:
    s = (s or '').strip().lower()
    s = re.sub(r"[\u00A0\s]+", " ", s)  # normaliza espaços e NBSP
    s = re.sub(r"[áàâã]", "a", s)
    s = re.sub(r"[éèê]", "e", s)
    s = re.sub(r"[íìî]", "i", s)
    s = re.sub(r"[óòôõ]", "o", s)
    s = re.sub(r"[úùû]", "u", s)
    s = re.sub(r"[^a-z0-9_]+", "_", s)
    return s.strip('_')

def _safe_int(x) -> int:
    try:
        if x is None: return 0
        if isinstance(x, (int,)): return int(x)
        s = str(x).strip().replace('\u00A0',' ')
        m = re.match(r"^[+\-]?\d+", s)
        return int(m.group(0)) if m else 0
    except Exception:
        return 0

def _parse_bonus_text(txt: str) -> Dict[str, int]:
    out: Dict[str, int] = {}
    if not txt:
        return out
    s = str(txt).replace('\u00A0', ' ')
    parts = [p.strip() for p in s.replace(';', ',').split(',') if p.strip()]
    for p in parts:
        m = re.search(r"([+\-]?\d+)\s*([A-Za-zçÇãÃéÉêÊíÍóÓúÚ]+)", p)
        if not m:
            continue
        val = _safe_int(m.group(1))
        label = _slug(m.group(2))
        k = ATTR_MAP.get(label)
        if k:
            out[k] = out.get(k, 0) + val
    return out

@dataclass
class Race:
    key: str
    name_pt: str
    name_en: str
    lore_pt: str
    lore_en: str
    attrs: Dict[str, int]
    perk: str

@dataclass
class Clazz:
    key: str
    name_pt: str
    name_en: str
    lore_pt: str
    lore_en: str
    attrs: Dict[str, int]
    focus: str

@dataclass
class Constellation:
    key: str
    name_pt: str
    name_en: str
    lore_pt: str
    lore_en: str
    bonus: Dict[str, int]
    perk: str

@dataclass
class Skill:
    key: str
    name_pt: str
    name_en: str
    desc_pt: str
    desc_en: str
    bonus: Dict[str, int]

_RACES: Dict[str, Race] = {}
_CLASSES: Dict[str, Clazz] = {}
_CONSTS: Dict[str, Constellation] = {}
_SKILLS: Dict[str, Skill] = {}

_RACE_BASES = {
    'planicius': {'HP': 60, 'MP': 30, 'STA': 50},
    'valcoran':  {'HP': 55, 'MP': 40, 'STA': 45},
    'solariar': {'HP': 40, 'MP': 70, 'STA': 40},
    'silvanor': {'HP': 45, 'MP': 35, 'STA': 65},
    'orvian':   {'HP': 70, 'MP': 25, 'STA': 50},
    'caelari':  {'HP': 45, 'MP': 35, 'STA': 60},
    'nerathi':  {'HP': 65, 'MP': 30, 'STA': 55},
}

def _lang() -> str:
    st = load_settings(); return st.get('language', 'en-US')

def _text(pt: str, en: str) -> str:
    return (en or pt or '').strip()

def _valid_attr_row(vals) -> bool:
    ints = [_safe_int(x) for x in vals]
    return any(v >= 0 for v in ints)  # permite 0; lista válida se colunas existem

def _load_sheet_races(ws):
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        name_pt = str(r[0]).strip()
        # ignora notas estranhas (nome muito longo, sem ser raça)
        if len(name_pt) > 60:
            continue
        attrs_vals = r[1:5]
        if len(attrs_vals) < 4:
            continue
        key = _slug(name_pt)
        attrs = { 'STR': _safe_int(r[1]), 'DEX': _safe_int(r[2]), 'INT': _safe_int(r[3]), 'CON': _safe_int(r[4]) }
        perk = str(r[5] or '').strip()
        lore_pt = str(r[6] or '').strip()
        _RACES[key] = Race(key, name_pt, name_pt, lore_pt, lore_pt, attrs, perk)

def _load_sheet_classes(ws):
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        name_pt = str(r[0]).strip(); key = _slug(name_pt)
        attrs = { 'STR': _safe_int(r[1]), 'DEX': _safe_int(r[2]), 'INT': _safe_int(r[3]), 'CON': _safe_int(r[4]) }
        focus = str(r[5] or '').strip(); lore_pt = str(r[6] or '').strip()
        _CLASSES[key] = Clazz(key, name_pt, name_pt, lore_pt, lore_pt, attrs, focus)

def _load_sheet_consts(ws):
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        name_pt = str(r[0]).strip(); key = _slug(name_pt)
        bonus = _parse_bonus_text(str(r[1] or ''))
        # Fallbacks específicos solicitados
        if not bonus:
            low = name_pt.strip().lower()
            if 'terradon' in low:
                bonus = {'CON': 2}
            elif 'vulkhar' in low:
                bonus = {'STR': 2}
        perk = str(r[2] or '').strip(); lore_pt = str(r[3] or '').strip()
        _CONSTS[key] = Constellation(key, name_pt, name_pt, lore_pt, lore_pt, bonus, perk)

def _load_sheet_skills(ws):
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        name_pt = str(r[0]).strip(); key = _slug(name_pt)
        bonus = _parse_bonus_text(str(r[1] or ''))
        # Fallbacks para Lâmina e Martelo
        if not bonus:
            low = name_pt.strip().lower()
            if 'lamina' in low or 'lâmina' in low:
                bonus = {'STR': 1}
            elif 'martelo' in low:
                bonus = {'STR': 2}
        desc_pt = str(r[2] or '').strip(); lore_pt = str(r[3] or '').strip()
        _SKILLS[key] = Skill(key, name_pt, name_pt, desc_pt, desc_pt, bonus)

def load_xlsx_once():
    if _RACES:
        return
    wb = load_workbook(str(XLSX_PATH))
    for sheet in wb.worksheets:
        try:
            header = [str(c or '').strip().lower() for c in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
            hdr = ' '.join(header)
        except StopIteration:
            header, hdr = [], ''
        title = (sheet.title or '').lower()
        if 'raça' in title or 'raca' in title or 'raça' in hdr or 'raca' in hdr:
            _load_sheet_races(sheet)
        elif 'classe' in title or 'classe' in hdr:
            _load_sheet_classes(sheet)
        elif 'constela' in title or 'constela' in hdr:
            _load_sheet_consts(sheet)
        elif 'perícia' in title or 'pericia' in title or 'perícia' in hdr or 'pericia' in hdr:
            _load_sheet_skills(sheet)

def current_lang() -> str: return _lang()

def race_keys() -> List[str]:
    load_xlsx_once(); return list(_RACES.keys())

def class_keys() -> List[str]:
    load_xlsx_once(); return list(_CLASSES.keys())

def const_keys() -> List[str]:
    load_xlsx_once(); return list(_CONSTS.keys())

def skill_keys() -> List[str]:
    load_xlsx_once(); return list(_SKILLS.keys())

def race_info(key: str) -> Race:
    load_xlsx_once(); return _RACES[key]

def class_info(key: str) -> Clazz:
    load_xlsx_once(); return _CLASSES[key]

def const_info(key: str) -> Constellation:
    load_xlsx_once(); return _CONSTS[key]

def skill_info(key: str) -> Skill:
    load_xlsx_once(); return _SKILLS[key]

def race_label(key: str) -> str:
    r = race_info(key); return r.name_pt if _lang().startswith('pt') else (r.name_en or r.name_pt)

def race_lore(key: str) -> str:
    r = race_info(key); return r.lore_pt if _lang().startswith('pt') else (r.lore_en or r.lore_pt)

def class_label(key: str) -> str:
    c = class_info(key); return c.name_pt if _lang().startswith('pt') else (c.name_en or c.name_pt)

def class_lore(key: str) -> str:
    c = class_info(key); return c.lore_pt if _lang().startswith('pt') else (c.lore_en or c.lore_pt)

def const_label(key: str) -> str:
    s = const_info(key); return s.name_pt if _lang().startswith('pt') else (s.name_en or s.name_pt)

def const_lore(key: str) -> str:
    s = const_info(key); return s.lore_pt if _lang().startswith('pt') else (s.lore_en or s.lore_pt)

def skill_label(key: str) -> str:
    sk = skill_info(key); return sk.name_pt if _lang().startswith('pt') else (sk.name_en or sk.name_pt)

def skill_desc(key: str) -> str:
    sk = skill_info(key); return sk.desc_pt if _lang().startswith('pt') else (sk.desc_en or sk.desc_pt)

def race_bases(key: str) -> Dict[str, int]:
    load_xlsx_once(); return _RACE_BASES.get(key, {'HP': 50, 'MP': 30, 'STA': 50})
